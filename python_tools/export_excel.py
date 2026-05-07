#!/usr/bin/env python3
# ════════════════════════════════════════════════
# COOP TAFERNOUT — Export vers Excel / CSV
# Exporte les données de l'application vers des
# fichiers CSV (et optionnellement Excel .xlsx)
# pour analyse dans un tableur (Excel, LibreOffice).
# ════════════════════════════════════════════════

import json
import os
import sys
import csv
from datetime import datetime


# ── Configuration ─────────────────────────────────────────────────────────────

DATA_FILE  = os.path.join(os.path.dirname(__file__), '..', 'tafernout_data.json')
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'exports')


# ── Chargement des données ─────────────────────────────────────────────────────

def load_data(path):
    if not os.path.exists(path):
        print(f"❌  Fichier introuvable : {path}")
        print("    Lancez d'abord l'application pour créer le fichier de données.")
        sys.exit(1)
    with open(path, 'r', encoding='utf-8-sig') as f:
        return json.load(f)


# ── Écriture CSV ───────────────────────────────────────────────────────────────

def write_csv(filepath, headers, rows):
    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f, delimiter=';')
        writer.writerow(headers)
        writer.writerows(rows)
    size = os.path.getsize(filepath)
    return size


# ── Export : Produits ──────────────────────────────────────────────────────────

def export_produits(data, dossier, date_str):
    produits = data.get('products', [])
    headers  = ['ID', 'Nom', 'Catégorie', 'Prix vente (MAD)', 'Coût revient (MAD)',
                 'Marge (%)', 'Stock actuel', 'Stock minimum', 'Alerte stock']
    rows = []
    for p in produits:
        marge  = round((p['price'] - p.get('cost', 0)) / p['price'] * 100, 1) if p['price'] else 0
        alerte = '⛔ Rupture' if p['stock'] <= 0 else ('⚠️ Faible' if p['stock'] <= p['minStock'] else '✅ OK')
        rows.append([
            p['id'], p['name'], p.get('category',''),
            p['price'], p.get('cost',0), marge,
            p['stock'], p.get('minStock',0), alerte
        ])
    path = os.path.join(dossier, f'produits_{date_str}.csv')
    sz   = write_csv(path, headers, rows)
    print(f"   ✅ produits_{date_str}.csv          ({len(rows):3d} lignes, {sz} octets)")
    return path


# ── Export : Ventes ────────────────────────────────────────────────────────────

def export_ventes(data, dossier, date_str):
    ventes  = data.get('sales', [])
    headers = ['ID', 'Date', 'Client', 'Articles', 'Quantités', 'Sous-total (MAD)',
                'Total (MAD)', 'Statut', 'Créance']
    rows = []
    for v in ventes:
        articles = ' | '.join(i['name'] for i in v.get('items', []))
        qtites   = ' | '.join(str(i['qty']) for i in v.get('items', []))
        rows.append([
            v['id'], v.get('date',''), v.get('client',''),
            articles, qtites,
            v.get('subtotal', v.get('total',0)),
            v.get('total',0),
            v.get('status',''),
            'Oui' if v.get('credit') else 'Non'
        ])
    path = os.path.join(dossier, f'ventes_{date_str}.csv')
    sz   = write_csv(path, headers, rows)
    print(f"   ✅ ventes_{date_str}.csv            ({len(rows):3d} lignes, {sz} octets)")
    return path


# ── Export : Clients ───────────────────────────────────────────────────────────

def export_clients(data, dossier, date_str):
    clients = data.get('clients', [])
    headers = ['ID', 'Nom', 'Téléphone', 'Créance (MAD)', 'Notes', 'Statut créance']
    rows    = []
    for c in clients:
        statut = '⚠️ Débiteur' if c.get('credit', 0) > 0 else '✅ Soldé'
        rows.append([
            c['id'], c['name'], c.get('phone',''),
            c.get('credit',0), c.get('notes',''), statut
        ])
    path = os.path.join(dossier, f'clients_{date_str}.csv')
    sz   = write_csv(path, headers, rows)
    print(f"   ✅ clients_{date_str}.csv           ({len(rows):3d} lignes, {sz} octets)")
    return path


# ── Export : Charges & Pertes ──────────────────────────────────────────────────

def export_charges(data, dossier, date_str):
    charges = data.get('charges', [])
    headers = ['ID', 'Date', 'Libellé', 'Catégorie', 'Type',
                'Montant (MAD)', 'Fournisseur', 'Payé']
    rows    = []
    for c in charges:
        rows.append([
            c['id'], c.get('date',''), c.get('label',''),
            c.get('category',''), c.get('type','charge'),
            c.get('amount',0), c.get('supplier',''),
            'Oui' if c.get('paid') else 'Non'
        ])
    path = os.path.join(dossier, f'charges_{date_str}.csv')
    sz   = write_csv(path, headers, rows)
    print(f"   ✅ charges_{date_str}.csv           ({len(rows):3d} lignes, {sz} octets)")
    return path


# ── Export : Employés ──────────────────────────────────────────────────────────

def export_employes(data, dossier, date_str):
    employes = data.get('employees', [])
    headers  = ['ID', 'Nom', 'Poste', 'Salaire mensuel (MAD)', 'Taux journalier (MAD)',
                 'Jours/mois', 'Téléphone', 'Date embauche', 'Statut']
    rows     = []
    for e in employes:
        rows.append([
            e['id'], e['name'], e.get('role',''),
            e.get('salary',0), e.get('dailyRate', round(e.get('salary',0)/26)),
            e.get('workDays',25), e.get('phone',''),
            e.get('startDate',''), e.get('status','actif')
        ])
    path = os.path.join(dossier, f'employes_{date_str}.csv')
    sz   = write_csv(path, headers, rows)
    print(f"   ✅ employes_{date_str}.csv          ({len(rows):3d} lignes, {sz} octets)")
    return path


# ── Export : Matières premières ────────────────────────────────────────────────

def export_matieres(data, dossier, date_str):
    matieres = data.get('rawMaterials', [])
    headers  = ['ID', 'Matière', 'Unité', 'Stock actuel', 'Stock minimum',
                 'Fournisseur', 'Dernière MAJ', 'Alerte']
    rows     = []
    for m in matieres:
        alerte = '⛔ Rupture' if m['currentStock'] <= 0 else ('⚠️ Faible' if m['currentStock'] <= m['minStock'] else '✅ OK')
        rows.append([
            m['id'], m['name'], m.get('unit',''),
            m['currentStock'], m.get('minStock',0),
            m.get('supplier',''), m.get('lastUpdated',''), alerte
        ])
    path = os.path.join(dossier, f'matieres_{date_str}.csv')
    sz   = write_csv(path, headers, rows)
    print(f"   ✅ matieres_{date_str}.csv          ({len(rows):3d} lignes, {sz} octets)")
    return path


# ── Export : Commandes ─────────────────────────────────────────────────────────

def export_commandes(data, dossier, date_str):
    commandes = data.get('commandes', [])
    headers   = ['N°', 'Date', 'Client', 'Téléphone', 'Articles', 'Total (MAD)', 'Livraison', 'Statut', 'Notes']
    rows      = []
    for cmd in commandes:
        articles = ' | '.join(f"{i['name']}×{i['qty']}" for i in cmd.get('items', []))
        rows.append([
            cmd.get('num',''), cmd.get('date',''), cmd.get('client',''),
            cmd.get('phone',''), articles, cmd.get('total',0),
            cmd.get('livraison',''), cmd.get('statut',''), cmd.get('notes','')
        ])
    path = os.path.join(dossier, f'commandes_{date_str}.csv')
    sz   = write_csv(path, headers, rows)
    print(f"   ✅ commandes_{date_str}.csv         ({len(rows):3d} lignes, {sz} octets)")
    return path


# ── Point d'entrée ─────────────────────────────────────────────────────────────

def main():
    data     = load_data(DATA_FILE)
    nom      = data.get('appSettings', {}).get('name', 'Coop Tafernout')
    date_str = datetime.now().strftime('%Y%m%d_%H%M')

    # Créer un sous-dossier horodaté pour cet export
    dossier  = os.path.join(OUTPUT_DIR, f"export_{date_str}")
    os.makedirs(dossier, exist_ok=True)

    print(f"\n🚀 Export CSV — {nom}")
    print(f"   Dossier : {dossier}")
    print("─"*55)

    fichiers = []
    fichiers.append(export_produits(data, dossier, date_str))
    fichiers.append(export_ventes(data, dossier, date_str))
    fichiers.append(export_clients(data, dossier, date_str))
    fichiers.append(export_charges(data, dossier, date_str))
    fichiers.append(export_employes(data, dossier, date_str))
    fichiers.append(export_matieres(data, dossier, date_str))
    fichiers.append(export_commandes(data, dossier, date_str))

    print("─"*55)
    print(f"✅ {len(fichiers)} fichiers exportés dans :")
    print(f"   {dossier}")

    # Tentative export Excel si openpyxl est disponible
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        xlsx_path = os.path.join(dossier, f'tafernout_complet_{date_str}.xlsx')
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Supprime la feuille vide par défaut

        header_font  = Font(bold=True, color='FFFFFF', size=9)
        header_fill  = PatternFill(fill_type='solid', fgColor='5C3317')
        header_align = Alignment(horizontal='center')

        # Charge et ajoute chaque CSV comme feuille Excel
        noms_feuilles = ['Produits', 'Ventes', 'Clients', 'Charges', 'Employés', 'Matières', 'Commandes']
        for i, (fpath, fnom) in enumerate(zip(fichiers, noms_feuilles)):
            ws = wb.create_sheet(title=fnom)
            with open(fpath, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f, delimiter=';')
                for row_num, row in enumerate(reader, 1):
                    for col_num, val in enumerate(row, 1):
                        cell = ws.cell(row=row_num, column=col_num, value=val)
                        if row_num == 1:
                            cell.font   = header_font
                            cell.fill   = header_fill
                            cell.alignment = header_align
            # Ajuste largeur colonnes
            for col in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        wb.save(xlsx_path)
        print(f"\n   📊 Fichier Excel généré : {os.path.basename(xlsx_path)}")
    except ImportError:
        print("\n   ℹ️  Pour générer un fichier Excel, installez openpyxl :")
        print("       pip install openpyxl")
    except Exception as e:
        print(f"\n   ⚠️  Impossible de créer le fichier Excel : {e}")

    print(f"\n   💡 Ouvrez les fichiers .csv avec Excel ou LibreOffice Calc.\n")


if __name__ == '__main__':
    main()
